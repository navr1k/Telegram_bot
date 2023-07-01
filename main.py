import telebot
from telebot import types
from openpyxl import load_workbook

wd_form = load_workbook(filename='остатки.xlsx', data_only=True)
sheet_form = wd_form['Лист1']

bot = telebot.TeleBot('6282442884:AAGB3mWmaes2DX73BWF2a6y7CnWRUqGMCwQ')
article = None
row = None
id_customer = None
id = 5583370934
check = False

@bot.message_handler(commands=['start'])
def main(message):
    global id_customer
    if message.from_user.last_name != None:
        bot.send_message(message.chat.id, f'Приветствую, {message.from_user.first_name} {message.from_user.last_name}! \nВведите артикул товара:')
        id_customer = message.from_user.username

    else:
        bot.send_message(message.chat.id, f'Приветствую, {message.from_user.first_name}! \nВведите артикул товара:')
        id_customer = message.from_user.username


@bot.message_handler()
def search(message):
    global row
    global article
    global id
    global id_customer
    global check
    if check:
        qty = message.text
        if qty.isnumeric():
            qty = int(qty)
            if qty >= 0 and sheet_form[row][2].value - qty >= 0:
                sheet_form[row][2].value = sheet_form[row][2].value - qty
                wd_form.save('остатки.xlsx')
                bot.send_message(message.chat.id,
                                 f'Ваш заказ {article} - {qty} штук успешно оформлен. \nДля нового заказа введите артикул товара:')
                bot.send_message(id,
                                 f'Заказ от {message.from_user.first_name} {message.from_user.last_name} (@{message.from_user.username}): {article} - {qty} штук.')
                check = False
            else:
                bot.send_message(message.chat.id, 'Неверное количество! Введите нужное количество.')
        else:
            bot.send_message(message.chat.id, 'Количество должно быть больше 0 и не должно содержать буквы! \nВведите нужное количество.')


    else:
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton('Заказать!', callback_data='qty'))
        for i in range(1, sheet_form.max_row):
            article = sheet_form[i][1].value[:-2]
            if message.text == article:
                row = i
                bot.reply_to(message, f'Остатков по артикулу: {article} - {sheet_form[i][2].value} штук.',
                             reply_markup=markup)
                break
        else:
            bot.send_message(message.chat.id, f'Артикул - {message.text} не найден!')



@bot.callback_query_handler(func=lambda callback: True)
def callback_message(callback):
    global check
    cancel = types.InlineKeyboardMarkup()
    cancel.add(types.InlineKeyboardButton('Отмена.', callback_data='cancel'))
    if callback.data == 'qty':
        bot.send_message(callback.from_user.id, 'Сколько рулонов вам нужно?', reply_markup=cancel)
        check = True
    elif callback.data == 'cancel':
        check = False
        bot.send_message(callback.from_user.id, 'Заказ отменён, введите новый артикул.')




bot.polling(none_stop=True)